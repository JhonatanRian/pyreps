from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rustpy_xlsxwriter import write_worksheet

from .contracts import OutputFormat, Renderer, ReportSpec
from .csv_options import CsvRenderOptions
from .xlsx_options import XlsxColumnOptions, XlsxRenderOptions


class CsvRenderer(Renderer):
    def render(
        self,
        rows: Iterable[Mapping[str, Any]],
        spec: ReportSpec,
        destination: str | Path,
    ) -> Path:
        output_path = Path(destination)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        options = CsvRenderOptions.from_metadata(spec.metadata)
        fieldnames = [column.label for column in spec.columns]
        with output_path.open("w", newline="", encoding=spec.encoding) as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=fieldnames,
                delimiter=options.delimiter,
                extrasaction="ignore",
            )
            writer.writeheader()
            writer.writerows(rows)
        return output_path


class XlsxRenderer(Renderer):
    def render(
        self,
        rows: Iterable[Mapping[str, Any]],
        spec: ReportSpec,
        destination: str | Path,
    ) -> Path:
        output_path = Path(destination)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        materialized_rows = list(rows)
        labels = [column.label for column in spec.columns]
        normalized_rows = [
            {label: row.get(label) for label in labels} for row in materialized_rows
        ]

        options = XlsxRenderOptions.from_metadata(spec.metadata)
        write_worksheet(normalized_rows, str(output_path), sheet_name=options.sheet_name)
        _apply_column_widths(output_path, labels, normalized_rows, options)
        return output_path


class PdfRenderer(Renderer):
    def render(
        self,
        rows: Iterable[Mapping[str, Any]],
        spec: ReportSpec,
        destination: str | Path,
    ) -> Path:
        output_path = Path(destination)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        labels = [column.label for column in spec.columns]
        materialized_rows = list(rows)

        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=landscape(A4),
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )

        header_row = [Paragraph(f"<b>{label}</b>", styles["Normal"]) for label in labels]
        data_rows = [
            [str("" if row.get(label) is None else row.get(label)) for label in labels]
            for row in materialized_rows
        ]
        paragraph_rows = [
            [Paragraph(cell, styles["Normal"]) for cell in row]
            for row in data_rows
        ]
        table_data = [header_row, *paragraph_rows]

        col_widths = _resolve_pdf_column_widths(labels, data_rows, doc.width)
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )

        doc.build([table, Spacer(1, 0.5 * cm)])
        return output_path


def default_renderer_registry() -> dict[OutputFormat, Renderer]:
    return {"csv": CsvRenderer(), "xlsx": XlsxRenderer(), "pdf": PdfRenderer()}


def _apply_column_widths(
    output_path: Path,
    labels: list[str],
    normalized_rows: list[dict[str, Any]],
    options: XlsxRenderOptions,
) -> None:
    workbook = load_workbook(output_path)
    worksheet = workbook[options.sheet_name]

    for index, label in enumerate(labels, start=1):
        values = [row.get(label) for row in normalized_rows]
        width = _resolve_width_for_label(label=label, values=values, options=options)
        worksheet.column_dimensions[get_column_letter(index)].width = width

    workbook.save(output_path)


def _resolve_width_for_label(
    *, label: str, values: list[Any], options: XlsxRenderOptions
) -> float:
    column_options = options.columns.get(label, XlsxColumnOptions())
    explicit = column_options.width

    if explicit is not None:
        width = explicit
    elif options.width_mode in {"auto", "mixed"}:
        max_len = max(
            [len(label), *(len("" if value is None else str(value)) for value in values)]
        )
        width = float(max_len) + options.auto_padding
    else:
        width = options.default_width

    if column_options.min_width is not None:
        width = max(width, column_options.min_width)
    if column_options.max_width is not None:
        width = min(width, column_options.max_width)

    return width


def _resolve_pdf_column_widths(
    labels: list[str],
    data_rows: list[list[str]],
    available_width: float,
) -> list[float]:
    """Distribute available page width proportionally based on max content length per column.

    Content length is capped to avoid a single column monopolising the page.
    Each column also gets a minimum width to ensure padding fits.
    """
    if not labels:
        return []

    _MAX_CHARS = 80
    _MIN_WIDTH_PT = 30.0  # enough to hold padding (8+8) plus a few characters

    char_widths = []
    for i, label in enumerate(labels):
        max_chars = len(label)
        for row in data_rows:
            if i < len(row):
                max_chars = max(max_chars, len(row[i]))
        char_widths.append(min(max(max_chars, 1), _MAX_CHARS))

    total_chars = sum(char_widths)
    widths = [available_width * (w / total_chars) for w in char_widths]

    # enforce minimum; redistribute surplus proportionally
    deficit = sum(max(0.0, _MIN_WIDTH_PT - w) for w in widths)
    if deficit > 0:
        surplus_indices = [i for i, w in enumerate(widths) if w > _MIN_WIDTH_PT]
        surplus_total = sum(widths[i] - _MIN_WIDTH_PT for i in surplus_indices)
        for i, w in enumerate(widths):
            if w < _MIN_WIDTH_PT:
                widths[i] = _MIN_WIDTH_PT
            elif surplus_total > 0:
                share = (w - _MIN_WIDTH_PT) / surplus_total
                widths[i] = _MIN_WIDTH_PT + (w - _MIN_WIDTH_PT) - deficit * share

    return widths
